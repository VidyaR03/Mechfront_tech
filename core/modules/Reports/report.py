from django.shortcuts import render, redirect, get_object_or_404
from core.models import  Performa_Invoice,Debit_Notes,Credit_Notes, customer,Purchase_Invoice,invoice_items, vendor, cash_voucher,sales_order_items, Invoice, inventory, Goods_received_notes,sales_order,Purchase_Invoice_items
from core.modules import template_path
from datetime import datetime
from django.http import JsonResponse
from django.http import HttpResponse,HttpResponseRedirect
from django.contrib import messages
from core.modules.Account.forms import DateRangeForm
import datetime
from django.db.models import Sum, F, ExpressionWrapper, DurationField, Func,IntegerField
from django.utils.timezone import now

from datetime import datetime
from core.modules.login.login import login_required
from django.urls import reverse
import openpyxl
from openpyxl.styles import Font
from django.utils.dateparse import parse_date
from io import BytesIO
import pandas as pd
from openpyxl.styles import Font, Alignment
from django.utils import timezone
from openpyxl.utils import get_column_letter
from collections import defaultdict
from django import template

import csv
from openpyxl import Workbook
from datetime import date
from dateutil.parser import parse
from django.utils import timezone


        
class Abs(Func):
    function = 'ABS'
    template = '%(function)s(%(expressions)s)'




@login_required
def debit_note_report(request):
    context = {
        'start_date': None,
        'end_date': None,
        'items_within_range': None,
        'filtered_items': None,
    }
 
    if request.method == 'POST':
        start_date_str = request.POST.get('start_date')
    
        end_date_str = request.POST.get('end_date')

        start_date = datetime.strptime(start_date_str, '%Y-%m-%d').date()
        end_date = datetime.strptime(end_date_str, '%Y-%m-%d').date()

        request.session['start_date'] = start_date_str
        request.session['end_date'] = end_date_str

        # Filter items within the date range
        items_within_range = Debit_Notes.objects.filter(cn_date__range=(start_date, end_date))
        filtered_items = list(items_within_range.values(
        'cn_customer_name', 'cn_date', 'cn_invoice_no', 'cn_total','cn_sub_total','cn_igstval','cn_sgstval','cn_cgstval'))
        total_amount = sum(float(item.cn_total or 0) for item in items_within_range)


       
        context.update({
            'items_within_range': items_within_range,
            'filtered_items': filtered_items,
            'start_date': start_date_str,
            'end_date': end_date_str,
            'total_amount':total_amount
        })

       
        return render(request, template_path.rdebit_note, context)
   

    return render(request,template_path.debit_note_date_report,context)


@login_required
def debit_note_date(request):
    return render(request, template_path.debit_note_date_report)
 





@login_required
def credit_note_report(request):
   
    context = {
        'start_date': None,
        'end_date': None,
        'items_within_range': None,
        'filtered_items': None,
    }
 
    if request.method == 'POST':
        start_date_str = request.POST.get('start_date')
    
        end_date_str = request.POST.get('end_date')

        start_date = datetime.strptime(start_date_str, '%Y-%m-%d').date()
        end_date = datetime.strptime(end_date_str, '%Y-%m-%d').date()

        request.session['start_date'] = start_date_str
        request.session['end_date'] = end_date_str

        # Filter items within the date range
        items_within_range = Credit_Notes.objects.filter(cn_date__range=(start_date, end_date))
        filtered_items = list(items_within_range.values(
        'cn_customer_name', 'cn_date', 'cn_invoice_no', 'cn_total','cn_sub_total','cn_igstval','cn_sgstval','cn_cgstval'))
        total_amount = sum(float(item.cn_total or 0) for item in items_within_range)


       
        context.update({
            'items_within_range': items_within_range,
            'filtered_items': filtered_items,
            'start_date': start_date_str,
            'end_date': end_date_str,
            'total_amount':total_amount
        })

       
        return render(request, template_path.rcredit_note, context)
   

    return render(request,template_path.credit_note_date_report,context)


@login_required
def credit_note_date(request):
    return render(request, template_path.credit_note_date_report)
 


@login_required
def download_credit_note_excel(request):
    start_date_str = request.session.get('start_date')
    end_date_str = request.session.get('end_date')

    # Validate date strings
    if not start_date_str or not end_date_str:
        return HttpResponse("Start date and end date are required.", status=400)

    try:
        start_date = datetime.strptime(start_date_str, '%Y-%m-%d').date()
        end_date = datetime.strptime(end_date_str, '%Y-%m-%d').date()
    except ValueError:
        return HttpResponse("Invalid date format. Please use YYYY-MM-DD.", status=400)

    # Retrieve data from the database based on the date range
    items_within_range = Credit_Notes.objects.filter(
        cn_date__range=(start_date, end_date)
    ).order_by('cn_date')

    # Extract specific fields for display
    filtered_items = list(items_within_range.values(
        'cn_customer_name__company_name', 'id','cn_date', 'cn_invoice_no','cn_cgstval',
          'cn_sgstval', 'cn_igstval','cn_sub_total','cn_total',
    ))

    # Calculate the total amount
    total_amount = sum(float(item['cn_total']) for item in filtered_items)

    # Create an in-memory output file for the new workbook
    output = BytesIO()
    workbook = openpyxl.Workbook()
    sheet = workbook.active

    # Add title and date range
    title = "M/S KEYMECH TECHNOLOGIES"
    subtitle = "Credit Note Summary Report"
    date_range = f"From {start_date_str} To {end_date_str}"

    sheet.merge_cells('A1:I1')
    sheet.merge_cells('A2:I2')
    sheet.merge_cells('A3:I3')

    title_cell = sheet.cell(row=1, column=1)
    subtitle_cell = sheet.cell(row=2, column=1)
    date_range_cell = sheet.cell(row=3, column=1)

    title_cell.value = title
    subtitle_cell.value = subtitle
    date_range_cell.value = date_range

    title_cell.font = Font(size=14, bold=True)
    subtitle_cell.font = Font(size=12, bold=True)
    date_range_cell.font = Font(size=10, bold=True)

    # Write headers
    headers = ['Customer Name','Credit Note No','Credit Note Date', 'Invoice No','CGST Value', 'SGST Value', 'IGST Value', ' Sub Total', ' Total', ]
    sheet.append([])
    sheet.append(headers)

    # Write data rows
    for item in filtered_items:
        sheet.append([
            item['cn_customer_name__company_name'],
            item['id'],
            item['cn_date'],
            item['cn_invoice_no'],
            item['cn_cgstval'],
            item['cn_sgstval'],
            item['cn_igstval'],
            item['cn_sub_total'],

            item['cn_total'],
        ])

    # Write total row
    sheet.append([])
    sheet.append(['', '', '', '','','','','Total',total_amount])

    column_widths = [ 15, 15, 15, 20, 15, 15, 15, 10,10]
    for i, column_width in enumerate(column_widths, start=1):
        sheet.column_dimensions[openpyxl.utils.get_column_letter(i)].width = column_width

    # Save the workbook to the output file
    workbook.save(output)
    output.seek(0)

    # Create the response
    response = HttpResponse(
        output,
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename=credit_note_summary.xlsx'
    return response




@login_required
def download_debit_note_excel(request):
    start_date_str = request.session.get('start_date')
    end_date_str = request.session.get('end_date')

    # Validate date strings
    if not start_date_str or not end_date_str:
        return HttpResponse("Start date and end date are required.", status=400)

    try:
        start_date = datetime.strptime(start_date_str, '%Y-%m-%d').date()
        end_date = datetime.strptime(end_date_str, '%Y-%m-%d').date()
    except ValueError:
        return HttpResponse("Invalid date format. Please use YYYY-MM-DD.", status=400)

    # Retrieve data from the database based on the date range
    items_within_range = Debit_Notes.objects.filter(
        cn_date__range=(start_date, end_date)
    ).order_by('cn_date')

    # Extract specific fields for display
    filtered_items = list(items_within_range.values(
        'cn_customer_name__company_name','id', 'cn_date', 'cn_invoice_no','cn_cgstval',
          'cn_sgstval','cn_igstval', 'cn_sub_total','cn_total',
    ))

    # Calculate the total amount
    total_amount = sum(float(item['cn_total']) for item in filtered_items)

    # Create an in-memory output file for the new workbook
    output = BytesIO()
    workbook = openpyxl.Workbook()
    sheet = workbook.active

    # Add title and date range
    title = "M/S KEYMECH TECHNOLOGIES"
    subtitle = "Debit Note Summary Report"
    date_range = f"From {start_date_str} To {end_date_str}"

    sheet.merge_cells('A1:I1')
    sheet.merge_cells('A2:I2')
    sheet.merge_cells('A3:I3')

    title_cell = sheet.cell(row=1, column=1)
    subtitle_cell = sheet.cell(row=2, column=1)
    date_range_cell = sheet.cell(row=3, column=1)

    title_cell.value = title
    subtitle_cell.value = subtitle
    date_range_cell.value = date_range

    title_cell.font = Font(size=14, bold=True)
    subtitle_cell.font = Font(size=12, bold=True)
    date_range_cell.font = Font(size=10, bold=True)

    # Write headers
    headers = [ 'Customer Name','Debit Note No','Date', 'Invoice No','CGST Value',  'SGST Value', 'IGST Value', ' Sub Total', ' Total', ]
    sheet.append([])
    sheet.append(headers)

    # Write data rows
    for item in filtered_items:
        sheet.append([
            item['cn_customer_name__company_name'],
            item['id'],
            item['cn_date'],
            item['cn_invoice_no'],
            item['cn_cgstval'],
            item['cn_sgstval'],
            item['cn_igstval'],
            item['cn_sub_total'],

            item['cn_total'],
        ])

    # Write total row
    sheet.append([])
    sheet.append(['', '', '', '','','','','Total',total_amount])

    # Set column widths for better visibility
    column_widths = [ 15, 15, 15, 20, 15, 15, 15, 10,10]
    for i, column_width in enumerate(column_widths, start=1):
        sheet.column_dimensions[openpyxl.utils.get_column_letter(i)].width = column_width

    # Save the workbook to the output file
    workbook.save(output)
    output.seek(0)

    # Create the response
    response = HttpResponse(
        output,
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename=debit_note_summary.xlsx'
    return response


@login_required
def inventory_summery(request):  
    companies = vendor.objects.values('company_name').distinct().order_by('company_name')

    return render(request, template_path.inventory_sum, {
        'company': companies,
    })


#  Sales Register
@login_required
def sales_register_date(request):

   
    return render(request, template_path.sales_register_date_report)



from django.http import JsonResponse
from django.views.decorators.csrf import csrf_exempt

@login_required
def sales_register_remark_ajax(request, id):
    if request.method == 'POST':
        record = get_object_or_404(Invoice, id=id)
        query = request.POST.get('query')
        record.inv_remark = query
        record.save()
        return JsonResponse({'status': 'success'})
    return JsonResponse({'status': 'fail'}, status=400)


@login_required
def display_sales_register_date_range(request):
    if request.method == 'POST':
        start_date_str = request.POST.get('start_date')
        end_date_str = request.POST.get('end_date')

        start_date = datetime.strptime(start_date_str, '%Y-%m-%d').date()
        end_date = datetime.strptime(end_date_str, '%Y-%m-%d').date()


        # Store the date range in the session
        request.session['start_date'] = start_date_str
        request.session['end_date'] = end_date_str

        # Filter items within the date range
        sales_register_within_range = Invoice.objects.filter(invoice_date__range=(start_date, end_date))
     
        # Calculate totals
        total_invoice = sum(float(item.invoice_due or 0) for item in sales_register_within_range)
        total_cgst = sum(float(item.invoice_cgstval or 0) for item in sales_register_within_range)
        total_sgst = sum(float(item.invoice_sgstval or 0) for item in sales_register_within_range)
        total_igst = sum(float(item.invoice_igstval or 0) for item in sales_register_within_range)
        total_amount = sum(float(item.invoice_total or 0) for item in sales_register_within_range)

        context = {
            'sales_register_within_range': sales_register_within_range,
            'total_cgst': total_cgst,
            'total_sgst': total_sgst,
            'total_igst': total_igst,
            'total_amount': total_amount,
            'start_date': start_date_str,
            'end_date': end_date_str,
            'total_invoice':total_invoice
        }
        return render(request, template_path.sales_register_list_report, context)
    
    return render(request, template_path.sales_register_date_report, {})





@login_required
def purchase_register_date(request):
   
    return render(request, template_path.purchase_register_date_report)




#************** CUSTOMER OUTSTANDING **************#

def customer_outstanding(request):
    cust = customer.objects.all()
    context = {
        'cust': cust,
    }
    return render(request, template_path.cust_out, context)

def customer_oustanding_date(request):
    if request.method == 'POST':
        start_date_str = request.POST.get('start_date')
        end_date_str = request.POST.get('end_date')
        customer_name = request.POST.get('customer_name')

        # Validate presence of date strings
        if not start_date_str or not end_date_str:
            return HttpResponse("Start date and end date are required.", status=400)

        try:
            start_date = datetime.strptime(start_date_str, '%Y-%m-%d').date()
            end_date = datetime.strptime(end_date_str, '%Y-%m-%d').date()
        except ValueError:
            return HttpResponse("Invalid date format. Please use YYYY-MM-DD.", status=400)

        # Save the date range and selected customer to the session
        request.session['start_date'] = start_date_str
        request.session['end_date'] = end_date_str
        request.session['customer_name'] = customer_name

        # Define a small tolerance value for calculations
        EPSILON = 1e-6

        # Filter invoices based on the date range and customer selection
        items_within_range = Invoice.objects.filter(invoice_date__range=(start_date, end_date))

        # Initialize totals for invoice and due amounts
        total_invoice_amount = 0
        total_due_amount = 0

        if customer_name == 'All':
            # Aggregating totals for "All" customers
            items_within_range = items_within_range.values('invoice_customer_name_customer__company_name').annotate(
                total_invoice_amount=Sum('invoice_total'),
                total_due_amount=Sum('invoice_due')
            )

            # Calculate totals for all customers, adding only positive amounts
            total_invoice_amount = items_within_range.aggregate(Sum('total_invoice_amount'))['total_invoice_amount__sum'] or 0
            total_due_amount = items_within_range.aggregate(Sum('total_due_amount'))['total_due_amount__sum'] or 0
        else:
            # Filtering invoices for a specific customer
            items_within_range = items_within_range.filter(invoice_customer_name_customer__customer=customer_name)
            current_date = timezone.now().date()

            # Annotate with due days for overdue calculation
            items_within_range = items_within_range.annotate(
                due_days=ExpressionWrapper(current_date - F('invoice_due_date'), output_field=DurationField())
            )

            # Calculate totals for a specific customer, adding only positive amounts
            total_invoice_amount = items_within_range.aggregate(Sum('invoice_total'))['invoice_total__sum'] or 0
            total_due_amount = items_within_range.aggregate(Sum('invoice_due'))['invoice_due__sum'] or 0

        # Ensure that only positive values are added to the total
        total_invoice_amount = max(total_invoice_amount, 0)
        total_due_amount = max(total_due_amount, 0)

        # Create the context for rendering the template
        context = {
            'sales_register_within_range': items_within_range,
            'total_invoice_amount': total_invoice_amount,
            'total_due_amount': total_due_amount,
            'selected_customer': customer_name,
            'start_date_str': start_date_str,
            'end_date_str': end_date_str,
            'start_date': start_date.strftime('%d-%m-%Y'),
            'end_date': end_date.strftime('%d-%m-%Y')
        }

        return render(request, template_path.cust_out_date, context)

    # If not a POST request, render the page with an empty context
    return render(request, template_path.cust_out_date)


def download_excel_customer_out(request):
    if request.method == 'POST':
        start_date_str = request.POST.get('start_date')
        end_date_str = request.POST.get('end_date')
        customer_name = request.POST.get('customer_name')

        if not start_date_str or not end_date_str:
            return HttpResponse("Start date and end date are required.", status=400)

        try:
            start_date = datetime.strptime(start_date_str, '%Y-%m-%d').date()
            end_date = datetime.strptime(end_date_str, '%Y-%m-%d').date()
        except ValueError:
            return HttpResponse("Invalid date format. Please use YYYY-MM-DD.", status=400)

        # Define a small tolerance value
        EPSILON = 1e-6

        # Filter invoices based on date range and customer selection
        items_within_range = Invoice.objects.filter(invoice_date__range=(start_date, end_date))

        if customer_name == 'All':
            items_within_range = items_within_range.values('invoice_customer_name_customer__company_name').annotate(
                total_invoice_amount=Sum('invoice_total'),
                total_due_amount=Sum('invoice_due')
            )  # Skip entries with 0 due amount
        else:
            items_within_range = items_within_range.filter(invoice_customer_name_customer__customer=customer_name)  # Skip entries with 0 due amount
            current_date = now().date()
            items_within_range = items_within_range.annotate(
                due_days=ExpressionWrapper(Abs(current_date - F('invoice_due_date')), output_field=DurationField())
            )

        # Create an Excel workbook
        output = BytesIO()
        workbook = openpyxl.Workbook()
        sheet = workbook.active

        # **Add Title, Subtitle, and Date Range**
        title = "M/S KEYMECH TECHNOLOGIES"
        subtitle = f"Customer Outstanding Report - {customer_name}"
        date_range = f"From {start_date.strftime('%d-%m-%Y')} To {end_date.strftime('%d-%m-%Y')}"

        sheet.merge_cells('A1:I1')
        sheet.merge_cells('A2:I2')
        sheet.merge_cells('A3:I3')

        title_cell = sheet.cell(row=1, column=1)
        subtitle_cell = sheet.cell(row=2, column=1)
        date_range_cell = sheet.cell(row=3, column=1)

        title_cell.value = title
        subtitle_cell.value = subtitle
        date_range_cell.value = date_range

        title_cell.font = Font(size=14, bold=True)
        subtitle_cell.font = Font(size=12, bold=True)
        date_range_cell.font = Font(size=10, bold=True)

        # Start headers from row 5
        if customer_name == 'All':
            headers = ['Customer Name', 'Total Invoice Amount', 'Total Due Amount']
        else:
            headers = ['Invoice No', 'Invoice Date', 'PO No', 'Payment Term', 'Due Date', 
                       'Invoice Amount', 'Due Amount', 'Due Days']

        sheet.append([])
        sheet.append(headers)

        for col in range(1, len(headers) + 1):
            sheet.cell(row=5, column=col).font = Font(bold=True)

        total_invoice_amount = 0
        total_due_amount = 0

        for item in items_within_range:
            if customer_name == 'All':
                # Ensure amounts are not negative
                total_invoice_amount += max(item['total_invoice_amount'] or 0, 0)
                total_due_amount += max(item['total_due_amount'] or 0, 0)

                sheet.append([
                    item['invoice_customer_name_customer__company_name'], 
                    max(item['total_invoice_amount'], 0),  # Ensure no negative values
                    max(item['total_due_amount'], 0)  # Ensure no negative values
                ])
            else:
                due_days = (now().date() - item.invoice_due_date).days
                invoice_amount = round(float(item.invoice_total or 0.0), 2)
                due_amount = round(float(item.invoice_due or 0.0), 2)

                # Ensure due amount is not negative
                due_amount = max(due_amount, 0)  # Correct any negative due amount

                # Ensure amounts are not negative
                total_invoice_amount += max(invoice_amount, 0)
                total_due_amount += due_amount  # Use the corrected due_amount

                sheet.append([
                    item.inv_number,
                    item.invoice_date.strftime('%d-%m-%Y'), 
                    item.invoice_buyer_order_no,
                    item.invoice_payment_terms, 
                    item.invoice_due_date.strftime('%d-%m-%Y'),
                    invoice_amount,  # Ensure no negative values
                    due_amount,  # Ensure no negative values
                    due_days
                ])

        # **Add Totals at the Bottom**
        sheet.append([])
        if customer_name == 'All':
            sheet.append(["TOTAL", total_invoice_amount, total_due_amount])
        else:
            sheet.append(["", "", "", "", "TOTAL", total_invoice_amount, total_due_amount])

        total_row = sheet.max_row
        sheet.cell(row=total_row, column=1).font = Font(bold=True)
        sheet.cell(row=total_row, column=2).font = Font(bold=True)
        sheet.cell(row=total_row, column=3).font = Font(bold=True)

        # **Ensure totals are not negative**
        total_invoice_amount = max(total_invoice_amount, 0)
        total_due_amount = max(total_due_amount, 0)

        # Set column widths
        for column_cells in sheet.columns:
            max_length = 0
            column_letter = None  # Initialize column letter

            for cell in column_cells:
                if cell.value and not isinstance(cell, openpyxl.cell.MergedCell):  # Skip merged cells
                    max_length = max(max_length, len(str(cell.value)))
                    column_letter = cell.column_letter  # Get column letter

            if column_letter:  # Ensure column_letter is set
                sheet.column_dimensions[column_letter].width = max_length + 2

        # Save the workbook to the output file
        workbook.save(output)
        output.seek(0)

        # Create the response
        response = HttpResponse(
            output,
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename=Customer_Outstanding_Report_{start_date}_to_{end_date}.xlsx'
        return response

    return HttpResponse("Invalid request method.", status=400)


#************** CUSTOMER OUTSTANDING END **************#



@login_required
def vendor_outstanding(request):
    vendor_out = vendor.objects.filter(purchase_invoice__isnull=False).distinct()
    total_amount = sum(
        float(t.receive_amount) for t in vendor_out
        if t.receive_amount and isinstance(t.receive_amount, (int, float))
    ) or 0.0
    
    context = {
        'vendor_out': vendor_out,
        'total_amount': total_amount
    }
    
    if request.method == "POST":
        customer_nm = request.POST.get('vendor_id')
        start_date_str = request.POST.get('start_date')
        end_date_str = request.POST.get('end_date')
        download_excel = request.POST.get('download_excel')


        # Validate inputs
        if not all([customer_nm, start_date_str, end_date_str]):
            return render(request, template_path.vendor_outs_report, {
                'error_message': 'Missing vendor or date inputs.',
                'selected_customer': customer_nm or 'all',
                'start_date': start_date_str,
                'end_date': end_date_str
            })

        # Validate date format
        try:
            datetime.strptime(start_date_str, '%Y-%m-%d')
            datetime.strptime(end_date_str, '%Y-%m-%d')
        except ValueError as e:
            return render(request, template_path.vendor_outs_report, {
                'error_message': 'Invalid date format. Please use YYYY-MM-DD.',
                'selected_customer': customer_nm or 'all',
                'start_date': start_date_str,
                'end_date': end_date_str
            })

        try:
            start_date = datetime.strptime(start_date_str, '%Y-%m-%d').date()
            end_date = datetime.strptime(end_date_str, '%Y-%m-%d').date()

            if start_date > end_date:
                return render(request, template_path.vendor_outs_report, {
                    'error_message': 'End date must be after start date.',
                    'selected_customer': customer_nm,
                    'start_date': start_date_str,
                    'end_date': end_date_str
                })

            # Filter Purchase_Invoice
            if customer_nm != 'all':
                try:
                    customer_detail = vendor.objects.get(id=customer_nm)
                    invoices = Purchase_Invoice.objects.filter(
                        Q(purchase_invoice_vendor_name=customer_nm) &
                        Q(purchase_invoice_date__range=[start_date, end_date])
                    ).select_related('purchase_invoice_vendor_name')
                except vendor.DoesNotExist:
                    return render(request, template_path.vendor_outs_report, {
                        'error_message': 'Selected vendor does not exist.',
                        'selected_customer': customer_nm,
                        'start_date': start_date_str,
                        'end_date': end_date_str
                    })
            else:
                invoices = Purchase_Invoice.objects.filter(
                    Q(purchase_invoice_date__range=[start_date, end_date])
                ).select_related('purchase_invoice_vendor_name')
                customer_detail = ''

         
            if not invoices.exists():
                return render(request, template_path.vendor_outs_report, {
                    'error_message': 'No data found for the selected date range.',
                    'selected_customer': customer_nm,
                    'start_date': start_date_str,
                    'end_date': end_date_str
                })

            total_amount_main = 0.0
            ven_entries = []
            vendor_totals = defaultdict(lambda: {"total_amount": 0.0})

            # Prepare data
            for invoice in invoices:
                try:
                    total_amount = float(invoice.purchase_invoice_total) if invoice.purchase_invoice_total and invoice.purchase_invoice_total.replace('.', '', 1).isdigit() else 0.0
                    total_amount_main += total_amount
                    ven_entries.append({
                        'invoice_number': invoice.id,
                        'date': invoice.purchase_invoice_date,
                        'amount': invoice.purchase_invoice_total,
                        'due_date': invoice.purchase_due_date,
                        'due_days': (invoice.purchase_due_date - invoice.purchase_invoice_date).days if invoice.purchase_due_date and invoice.purchase_invoice_date else 0,
                        'vendor_name': invoice.purchase_invoice_vendor_name.company_name if invoice.purchase_invoice_vendor_name else 'N/A'
                    })
                except Exception as e:
                    print(f"Error processing Purchase_Invoice {invoice.id}: {str(e)}")


            ven_entries.sort(key=lambda x: x['date'])

            # Group by vendor for 'all' case
            if customer_nm == 'all':
                for entry in ven_entries:
                    vendor_name = entry['vendor_name']
                    vendor_totals[vendor_name]["total_amount"] += float(entry['amount']) if entry['amount'] and entry['amount'].replace('.', '', 1).isdigit() else 0.0
                sorted_vendor_totals = dict(sorted(vendor_totals.items()))
            else:
                sorted_vendor_totals = {}

            if download_excel:
                # Create DataFrame
                df = pd.DataFrame(ven_entries)
                df['date'] = df['date'].apply(lambda x: x.strftime('%d %B %Y') if x else '')
                df['due_date'] = df['due_date'].apply(lambda x: x.strftime('%d %B %Y') if x else '')
                if customer_nm == 'all':
                    columns = ['vendor_name', 'invoice_number', 'date', 'due_date', 'amount', 'due_days']
                    df = df[columns]
                    df.insert(0, 'sr_no', range(1, len(df) + 1))
                    df.columns = ['SR.NO.', 'VENDOR NAME', 'INVOICE NO.', 'INVOICE DATE', 'DUE DATE', 'INVOICE AMOUNT', 'DUE DAYS']
                else:
                    columns = ['vendor_name', 'invoice_number', 'date', 'due_date', 'amount', 'due_days']
                    df = df[columns]
                    df.insert(0, 'sr_no', range(1, len(df) + 1))
                    df.columns = ['SR.NO.', 'VENDOR NAME', 'INVOICE NO.', 'INVOICE DATE', 'DUE DATE', 'INVOICE AMOUNT', 'DUE DAYS']

                # Add total row
                if customer_nm == 'all':
                    total_list = ['', '', '', 'TOTAL', '', total_amount_main, '']
                else:
                    total_list = ['', '', '', 'TOTAL', '', total_amount_main, '']
                total_row = pd.DataFrame([total_list], columns=df.columns)
                df = pd.concat([df, total_row], ignore_index=True)

                # Create Excel file
                output = BytesIO()
                with pd.ExcelWriter(output, engine='openpyxl') as writer:
                    df.to_excel(writer, index=False, sheet_name='Account Payable Summary', startrow=4)
                    sheet = writer.sheets['Account Payable Summary']

                    # Add title, subtitle, and date range
                    title = "M/S KEYMECH TECHNOLOGIES"
                    subtitle = "Account Payable Summary Details"
                    date_range = f"From {start_date_str} To {end_date_str}"

                    if customer_nm == 'all':
                        sheet.merge_cells('A1:H1')
                        sheet.merge_cells('A2:H2')
                        sheet.merge_cells('A3:H3')
                    else:
                        sheet.merge_cells('A1:H1')
                        sheet.merge_cells('A2:H2')
                        sheet.merge_cells('A3:H3')

                    title_cell = sheet.cell(row=1, column=1)
                    subtitle_cell = sheet.cell(row=2, column=1)
                    date_range_cell = sheet.cell(row=3, column=1)

                    title_cell.value = title
                    subtitle_cell.value = subtitle
                    date_range_cell.value = date_range

                    if customer_nm != 'all':
                        sheet.merge_cells('A4:G4')
                        vendor_cell = sheet.cell(row=4, column=1)
                        vendor_cell.value = customer_detail.company_name
                        vendor_cell.font = Font(size=12, bold=True)
                        vendor_cell.alignment = Alignment(horizontal='center')

                    title_cell.font = Font(size=14, bold=True)
                    subtitle_cell.font = Font(size=12, bold=True)
                    date_range_cell.font = Font(size=10, bold=True)

                    title_cell.alignment = Alignment(horizontal='center')
                    subtitle_cell.alignment = Alignment(horizontal='center')
                    date_range_cell.alignment = Alignment(horizontal='center')

                    # Style table headers
                    for cell in sheet[5]:
                        cell.font = Font(bold=True)
                        cell.alignment = Alignment(horizontal='center')

                    # Set fixed column widths
                    if customer_nm == 'all':
                        column_widths = [15, 15, 15, 15, 20, 15, 15, 15]
                    else:
                        column_widths = [15, 15, 15, 20, 15, 15, 15, 15]
                    for i, column_width in enumerate(column_widths, start=1):
                        sheet.column_dimensions[get_column_letter(i)].width = column_width

                output.seek(0)
                response = HttpResponse(
                    content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
                )
                response['Content-Disposition'] = f'attachment; filename=Account_Payable_Summary_{start_date_str}_to_{end_date_str}.xlsx'
                response.write(output.read())
                return response

            context = {
                'selected_customer': customer_nm,
                'ven_entries': ven_entries,
                'customer': customer_detail,
                'start_date': start_date_str,
                'end_date': end_date_str,
                'start_date_str': start_date.strftime('%d-%m-%Y') if start_date else '',
                'end_date_str': end_date.strftime('%d-%m-%Y') if end_date else '',
                'total_amount_main': total_amount_main,
                'vendor_totals': dict(sorted_vendor_totals)
            }
            return render(request, template_path.vendor_outs_report, context)

        except ValueError as e:
            return render(request, template_path.vendor_outs_report, {
                'error_message': 'Invalid date format. Please use YYYY-MM-DD.',
                'selected_customer': customer_nm or 'all',
                'start_date': start_date_str,
                'end_date': end_date_str
            })
        except vendor.DoesNotExist:
            return render(request, template_path.vendor_outs_report, {
                'error_message': 'Selected vendor does not exist.',
                'selected_customer': customer_nm,
                'start_date': start_date_str,
                'end_date': end_date_str
            })

    return render(request, template_path.vendor_outs, context)
@login_required
def expenses_report_date(request):
    return render(request, template_path.expense_date_wise)


@login_required
def sales_order_report_date(request):
   
    
    return render(request, template_path.sales_order_report_date)


@login_required
def vendor_order_report_date(request):
   
    
    return render(request, template_path.vendor_outs)





@login_required
def cash_voucher_date(request):
   
    return render(request, template_path.cash_voucher_date)



@login_required
def display_items_by_date_range(request):
    if request.method == 'POST':
        start_date_str = request.POST.get('start_date')
        end_date_str = request.POST.get('end_date')
        download_excel = request.POST.get('download_excel')  # Check if download button was clicked

        try:
            start_date = datetime.strptime(start_date_str, '%Y-%m-%d').date()
            end_date = datetime.strptime(end_date_str, '%Y-%m-%d').date()

            if start_date > end_date:
                return render(request, template_path.sales_order_date_report, {
                    'error_message': 'End date must be after start date.'
                })

            # Filter sales_order_items based on related sales_order's so_date
            items_within_range = sales_order_items.objects.filter(
                so_sales_order_id__in=sales_order.objects.filter(
                    so_date__range=(start_date, end_date)
                ).values_list('id', flat=True)
            )

            if not items_within_range.exists():
                return render(request, template_path.sales_order_date_report, {
                    'error_message': 'No data found for the selected date range.'
                })

            # List to store modified items
            modified_items = []

            for item in items_within_range:
                try:
                    so = sales_order.objects.get(id=item.so_sales_order_id)
                    customer = so.so_customer_name.company_name if so.so_customer_name else 'N/A'
                    buyer_ord_no = so.so_buyer_order_no or 'N/A'

                    modified_item = {
                        'sr_no': len(modified_items) + 1,
                        'date': so.so_date,
                        'company_name': customer,
                        'so_buyer_order_no': buyer_ord_no,
                        'so_description_goods': item.so_description_goods,
                        'so_qantity': item.so_qantity,
                        'so_uom': item.so_uom,
                    }
                    modified_items.append(modified_item)
                except sales_order.DoesNotExist:
                    pass

            # Calculate total quantity
            total_quantity = sum(float(item.so_qantity) for item in items_within_range if item.so_qantity and item.so_qantity.replace('.', '', 1).isdigit())

            if download_excel:
                # Create a DataFrame for Excel
                df = pd.DataFrame(modified_items)
                df['date'] = df['date'].apply(lambda x: x.strftime('%d %B %Y'))
                df = df[['sr_no', 'date', 'company_name', 'so_buyer_order_no', 'so_description_goods', 'so_qantity', 'so_uom']]
                df.columns = ['SR.NO.', 'DATE', 'COMPANY NAME', 'PO NO.', 'ITEM DESCRIPTION', 'QUANTITY', 'UOM']

                # Add total row
                total_row = pd.DataFrame([['', '', '', '', 'TOTAL', total_quantity, '']], columns=df.columns)
                df = pd.concat([df, total_row], ignore_index=True)

                # Create Excel file
                output = BytesIO()
                with pd.ExcelWriter(output, engine='openpyxl') as writer:
                    # Write DataFrame starting at row 5 to leave space for headers
                    df.to_excel(writer, index=False, sheet_name='Sales Order Report', startrow=4)

                    # Access the worksheet to add title, subtitle, and date range
                    sheet = writer.sheets['Sales Order Report']

                    # Add title, subtitle, and date range
                    title = "M/S KEYMECH TECHNOLOGIES"
                    subtitle = "Sales Orders Report"  # Adjusted to match context
                    date_range = f"From {start_date_str} To {end_date_str}"

                    sheet.merge_cells('A1:G1')  # Adjusted to 7 columns (A to G)
                    sheet.merge_cells('A2:G2')
                    sheet.merge_cells('A3:G3')

                    title_cell = sheet.cell(row=1, column=1)
                    subtitle_cell = sheet.cell(row=2, column=1)
                    date_range_cell = sheet.cell(row=3, column=1)

                    title_cell.value = title
                    subtitle_cell.value = subtitle
                    date_range_cell.value = date_range

                    title_cell.font = Font(size=14, bold=True)
                    subtitle_cell.font = Font(size=12, bold=True)
                    date_range_cell.font = Font(size=10, bold=True)

                    # Center align the header cells
                    title_cell.alignment = Alignment(horizontal='center')
                    subtitle_cell.alignment = Alignment(horizontal='center')
                    date_range_cell.alignment = Alignment(horizontal='center')

                    # Style the table headers
                    for cell in sheet[5]:  # Row 5 is the header row (1-based indexing)
                        cell.font = Font(bold=True)
                        cell.alignment = Alignment(horizontal='center')

                # Prepare response
                output.seek(0)
                response = HttpResponse(
                    content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
                )
                response['Content-Disposition'] = f'attachment; filename=Sales_Order_Report_{start_date_str}_to_{end_date_str}.xlsx'
                response.write(output.read())
                return response

            # Render the display page
            context = {
                'filtered_items': modified_items,
                'total_quantity': total_quantity,
                'start_date': start_date_str,
                'end_date': end_date_str
            }
            return render(request, template_path.sales_order_list_report, context)

        except ValueError:
            return render(request, template_path.sales_order_date_report, {
                'error_message': 'Invalid date format. Please use YYYY-MM-DD.'
            })

    return render(request, template_path.sales_order_date_report)

    
@login_required
def display_cash_voucher_by_date_range(request):
    start_date = None
    end_date = None
    if request.method == 'POST':
        start_date_str = request.POST.get('start_date')
        end_date_str = request.POST.get('end_date')

        start_date = datetime.strptime(start_date_str, '%Y-%m-%d').date()
        end_date = datetime.strptime(end_date_str, '%Y-%m-%d').date()

        # Filter items within the date range
        items_within_range = cash_voucher.objects.filter(date__range=(start_date, end_date))

        # Extract specific fields for display
        filtered_items = items_within_range.values(
            'date', 'bill_no', 'voucher_number', 'amount', 'particular', 'pay_to'
        )

        # Calculate total amount
        total_amount = items_within_range.aggregate(total_amount=Sum('amount'))['total_amount'] or 0

        return render(request, template_path.cash_voucher_list_report, {
            'filtered_items': filtered_items,
            'total_amount': total_amount ,
             'start_date': start_date,
        'end_date': end_date, # Include total_amount in the context
        })
   

    return render(request, template_path.cash_voucher_date_report)


@login_required
def display_invoice_by_date_range(request):
    if request.method == 'POST':
        start_date_str = request.POST.get('start_date')
        end_date_str = request.POST.get('end_date')
        download_excel = request.POST.get('download_excel')

        try:
            start_date = datetime.strptime(start_date_str, '%Y-%m-%d').date()
            end_date = datetime.strptime(end_date_str, '%Y-%m-%d').date()

            if start_date > end_date:
                return render(request, template_path.invoice_date_report, {
                    'error_message': 'End date must be after start date.'
                })

            # Filter invoice_items based on related Invoice's invoice_date
            invoice_ids = Invoice.objects.filter(
                invoice_date__range=(start_date, end_date)
            ).values_list('id', flat=True)

            items_within_range = invoice_items.objects.filter(
                invoice_id__in=invoice_ids
            )

            if not items_within_range.exists():
                return render(request, template_path.invoice_date_report, {
                    'error_message': 'No data found for the selected date range.'
                })

            # Pre-fetch Invoice objects
            invoice_dict = {
                str(invoice.id): invoice for invoice in Invoice.objects.filter(
                    id__in=invoice_ids
                ).select_related('invoice_customer_name_customer')
            }

            # List to store modified items
            modified_items = []

            for item in items_within_range:
                try:
                    invoice_id_str = str(item.invoice_id).strip()
                    inv = invoice_dict.get(invoice_id_str)
                    if not inv:
                        print(f"No Invoice found for invoice_id: {invoice_id_str}")
                        continue
                    customer = inv.invoice_customer_name_customer.customer if inv.invoice_customer_name_customer else 'N/A'
                    buyer_ord_no = inv.invoice_buyer_order_no or 'N/A'

                    modified_item = {
                        'sr_no': len(modified_items) + 1,
                        'date': inv.invoice_date,
                        'company_name': customer,
                        'invoice_buyer_order_no': buyer_ord_no,
                        'invoice_description_goods': item.invoice_description_goods,
                        'invoice_qantity': item.invoice_qantity,
                        'invoice_uom': item.invoice_uom,
                    }
                    modified_items.append(modified_item)
                except Exception as e:
                    print(f"Error processing invoice_item {item.id}: {str(e)}")


            # Calculate total quantity
            total_quantity = sum(
                float(item.invoice_qantity) for item in items_within_range 
                if item.invoice_qantity and item.invoice_qantity.replace('.', '', 1).isdigit()
            )

            if download_excel:
                df = pd.DataFrame(modified_items)
                df['date'] = df['date'].apply(lambda x: x.strftime('%d %B %Y'))
                df = df[['sr_no', 'date', 'company_name', 'invoice_buyer_order_no', 
                         'invoice_description_goods', 'invoice_qantity', 'invoice_uom']]
                df.columns = ['SR.NO.', 'DATE', 'CUSTOMER NAME', 'PO NO.', 'ITEM DESCRIPTION', 
                              'QUANTITY', 'UOM']

                total_row = pd.DataFrame([['', '', '', '', 'TOTAL', total_quantity, '']], 
                                        columns=df.columns)
                df = pd.concat([df, total_row], ignore_index=True)

                output = BytesIO()
                with pd.ExcelWriter(output, engine='openpyxl') as writer:
                    df.to_excel(writer, index=False, sheet_name='Invoice Report', startrow=4)
                    sheet = writer.sheets['Invoice Report']

                    title = "M/S KEYMECH TECHNOLOGIES"
                    subtitle = "Invoice Report"
                    date_range = f"From {start_date_str} To {end_date_str}"

                    sheet.merge_cells('A1:G1')
                    sheet.merge_cells('A2:G2')
                    sheet.merge_cells('A3:G3')

                    title_cell = sheet.cell(row=1, column=1)
                    subtitle_cell = sheet.cell(row=2, column=1)
                    date_range_cell = sheet.cell(row=3, column=1)

                    title_cell.value = title
                    subtitle_cell.value = subtitle
                    date_range_cell.value = date_range

                    title_cell.font = Font(size=14, bold=True)
                    subtitle_cell.font = Font(size=12, bold=True)
                    date_range_cell.font = Font(size=10, bold=True)

                    title_cell.alignment = Alignment(horizontal='center')
                    subtitle_cell.alignment = Alignment(horizontal='center')
                    date_range_cell.alignment = Alignment(horizontal='center')

                    for cell in sheet[5]:
                        cell.font = Font(bold=True)
                        cell.alignment = Alignment(horizontal='center')

                output.seek(0)
                response = HttpResponse(
                    content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
                )
                response['Content-Disposition'] = f'attachment; filename=Invoice_Report_{start_date_str}_to_{end_date_str}.xlsx'
                response.write(output.read())
                return response

            context = {
                'filtered_items': modified_items,
                'total_quantity': total_quantity,
                'start_date': start_date_str,
                'end_date': end_date_str
            }
            return render(request, template_path.invoice_list_report, context)

        except ValueError:
            return render(request, template_path.invoice_date_report, {
                'error_message': 'Invalid date format. Please use YYYY-MM-DD.'
            })

    return render(request, template_path.invoice_date_report)


def display_inventory_summenry_date_range(request):
    # Fetch all vendors, ordered by company_name
    vendors = vendor.objects.all().order_by('company_name')
    selected_company = ''
    filtered_items = []
    total_amount = 0
    start_date = None
    end_date = None

    if request.method == 'POST':
        start_date_str = request.POST.get('start_date')
        end_date_str = request.POST.get('end_date')
        selected_company = request.POST.get('company_name', '').strip()

        # Save to session for persistence
        request.session['start_date'] = start_date_str
        request.session['end_date'] = end_date_str
        request.session['company_name'] = selected_company

        # Convert date strings to date objects
        if start_date_str and end_date_str:
            try:
                start_date = datetime.strptime(start_date_str, '%Y-%m-%d').date()
                end_date = datetime.strptime(end_date_str, '%Y-%m-%d').date()

                # Filter inventory items by date range
                items_within_range = inventory.objects.filter(
                    date__range=(start_date, end_date)
                )

                # Filter by company if selected
                if selected_company:
                    items_within_range = items_within_range.filter(
                        vendor_name__company_name__icontains=selected_company
                    )

                # Annotate and order by item_code as integer
                items_within_range = items_within_range.annotate(
                    item_code_int=Cast('item_code', IntegerField())
                ).order_by('item_code_int')

                # Prepare result data
                filtered_items = list(items_within_range.values(
                    'item_code', 'vendor_name__company_name', 'inventory_name', 
                    'opening_stock_quantity','opening_rate', 'purchase_rate'
                ))

                # Calculate total amount
                for item in filtered_items:
                    try:
                        qty = float(item['opening_stock_quantity'])
                        rate = float(item['purchase_rate'])
                    except (ValueError, TypeError):
                        qty, rate = 0, 0
                    item['amount'] = qty * rate
                    total_amount += item['amount']
            except ValueError:
                # Handle invalid date format
                pass

    return render(request, template_path.inventory_summery_list_report, {
        'vendors': vendors,
        'selected_company': selected_company,
        'filtered_items': filtered_items,
        'total_amount': total_amount,
        'start_date': start_date,
        'end_date': end_date
    })

# from datetime import datetime
# from django.db.models import IntegerField
# from django.db.models.functions import Cast

# def display_inventory_summenry_date_range(request):
#     vendors = vendor.objects.all().order_by('company_name')
#     selected_company = ''
#     filtered_items = []
#     total_amount = 0
#     available_qty = 0
#     total_in_qty = 0
#     total_out_qty = 0

#     start_date = None
#     end_date = None

#     if request.method == 'POST':
#         start_date_str = request.POST.get('start_date')
#         end_date_str = request.POST.get('end_date')
#         selected_company = request.POST.get('company_name', '').strip()

#         request.session['start_date'] = start_date_str
#         request.session['end_date'] = end_date_str
#         request.session['company_name'] = selected_company

#         if start_date_str and end_date_str:
#             try:
#                 start_date = datetime.strptime(start_date_str, '%Y-%m-%d').date()
#                 end_date = datetime.strptime(end_date_str, '%Y-%m-%d').date()

#                 items_within_range = inventory.objects.filter(
#                     date__range=(start_date, end_date)
#                 )

#                 if selected_company:
#                     items_within_range = items_within_range.filter(
#                         vendor_name__company_name__icontains=selected_company
#                     )

#                 items_within_range = items_within_range.annotate(
#                     item_code_int=Cast('item_code', IntegerField())
#                 ).order_by('item_code_int')

#                 # Group by item_code to calculate running stock
#                 item_groups = {}

#                 for entry in items_within_range:
#                     code = entry.item_code

#                     if code not in item_groups:
#                         item_groups[code] = {
#                             'entries': [],
#                             'opening_stock_quantity': 0.0,
#                             'purchase_rate': 0.0,
#                             'total_in_qty': 0.0,
#                             'total_out_qty': 0.0
#                         }

#                     # Parse opening stock
#                     try:
#                         opening_stock = float(entry.opening_stock_quantity)
#                     except (ValueError, TypeError):
#                         opening_stock = 0.0

#                     item_groups[code]['opening_stock_quantity'] = opening_stock

#                     # Save purchase_rate just once (for amount calc)
#                     try:
#                         item_groups[code]['purchase_rate'] = float(entry.purchase_rate)
#                     except (ValueError, TypeError):
#                         item_groups[code]['purchase_rate'] = 0.0

#                     # Fetch sales data (OUT)
#                     invoices = Invoice.objects.filter(
#                         id__in=invoice_items.objects.filter(
#                             invoice_item_code=entry.item_code
#                         ).values_list('invoice_id', flat=True),
#                         invoice_date__range=(start_date, end_date)
#                     )
#                     invoice_items_data = invoice_items.objects.filter(
#                         invoice_item_code=entry.item_code,
#                         invoice_id__in=invoices.values_list('id', flat=True)
#                     )
#                     total_out = sum(float(item.invoice_qantity) for item in invoice_items_data)

#                     # Fetch purchase data (IN)
#                     purchase_invoices = Purchase_Invoice.objects.filter(
#                         id__in=Purchase_Invoice_items.objects.filter(
#                             purchase_invoice_item_code=entry.item_code
#                         ).values_list('purchase_invoice_id', flat=True),
#                         purchase_invoice_date__range=(start_date, end_date)
#                     )
#                     purchase_invoice_items_data = Purchase_Invoice_items.objects.filter(
#                         purchase_invoice_item_code=entry.item_code,
#                         purchase_invoice_id__in=purchase_invoices.values_list('id', flat=True)
#                     )
#                     total_in = sum(float(item.purchase_invoice_qantity) for item in purchase_invoice_items_data)

#                     # Store IN and OUT quantities
#                     item_groups[code]['total_in_qty'] = total_in
#                     item_groups[code]['total_out_qty'] = total_out
#                     item_groups[code]['entries'].append({
#                         'item_code': entry.item_code,
#                         'vendor_name__company_name': entry.vendor_name.company_name if entry.vendor_name else '',
#                         'inventory_name': entry.inventory_name,
#                         'opening_stock_quantity': opening_stock,
#                         'purchase_rate': entry.purchase_rate,
#                         'type': entry.type,
#                         'date': entry.date
#                     })

#                 processed_items = []

#                 for code, data in item_groups.items():
#                     balance_stock = data['opening_stock_quantity']
#                     in_qty = 0.0
#                     out_qty = 0.0

#                     for entry in sorted(data['entries'], key=lambda e: e.get('date')):
#                         try:
#                             qty = float(entry['opening_stock_quantity'])
#                         except (ValueError, TypeError):
#                             qty = 0.0

#                         if entry.get('type') == 'Purchase':
#                             balance_stock += qty
#                             in_qty += qty
#                         elif entry.get('type') == 'Sales':
#                             balance_stock -= qty
#                             out_qty += qty

#                         entry['balance'] = balance_stock

#                     amount = balance_stock * data['purchase_rate']
#                     total_amount += amount

#                     summary_entry = data['entries'][0]  # Use the first entry as base
#                     summary_entry['final_balance'] = balance_stock
#                     summary_entry['in_qty'] = in_qty       # ✅ Add In Qty
#                     summary_entry['out_qty'] = out_qty     # ✅ Add Out Qty
#                     summary_entry['amount'] = amount

#                     processed_items.append(summary_entry)


#                     total_amount += summary_entry['amount']
#                     total_in_qty += data['total_in_qty']
#                     total_out_qty += data['total_out_qty']

#                 filtered_items = processed_items
#                 available_qty = sum(item['final_balance'] for item in processed_items)

#             except ValueError:
#                 pass

#     return render(request, template_path.inventory_summery_list_report, {
#         'vendors': vendors,
#         'selected_company': selected_company,
#         'filtered_items': filtered_items,
#         'total_amount': total_amount,
#         'start_date': start_date,
#         'end_date': end_date,
#         'available_qty': available_qty,
#         'total_in_qty': total_in_qty,
#         'total_out_qty': total_out_qty
#     })
def get_companies_by_date(request):
    start_date = request.GET.get("start_date")
    end_date = request.GET.get("end_date")

    if not start_date or not end_date:
        return JsonResponse([], safe=False)

    try:
        start_date_obj = datetime.strptime(start_date, "%Y-%m-%d").date()
        end_date_obj = datetime.strptime(end_date, "%Y-%m-%d").date()
    except ValueError:
        return JsonResponse([], safe=False)

    # Get vendor names from inventory records in the date range
    companies = (
        inventory.objects.filter(date__range=(start_date_obj, end_date_obj))
        .values_list("vendor_name__company_name", flat=True)
        .distinct()
    )

    return JsonResponse(list(companies), safe=False)



@login_required
def display_purchase_register_date_range(request):
    if request.method == 'POST':
        start_date_str = request.POST.get('start_date')
        end_date_str = request.POST.get('end_date')

        start_date = datetime.strptime(start_date_str, '%Y-%m-%d').date()
        end_date = datetime.strptime(end_date_str, '%Y-%m-%d').date()

        # Filter items within the date range
        purchase_items_within_range = Purchase_Invoice.objects.filter(purchase_invoice_date__range=(start_date, end_date))

        # Save the date range in the session
        request.session['start_date'] = start_date_str
        request.session['end_date'] = end_date_str

        # Calculate totals
        total_cgst = sum(float(item.purchase_invoice_cgstval or 0) for item in purchase_items_within_range)
        total_sgst = sum(float(item.purchase_invoice_sgstval or 0) for item in purchase_items_within_range)
        total_igst = sum(float(item.purchase_invoice_igstval or 0) for item in purchase_items_within_range)
        total_amount = sum(float(item.purchase_invoice_sub_total or 0) for item in purchase_items_within_range)
        total_invoice = sum(float(item.purchase_invoice_total or 0) for item in purchase_items_within_range)
        total_freight = sum(float(item.purchase_freight_amount or 0) for item in purchase_items_within_range)

        context = {
            'purchase_items_within_range': purchase_items_within_range,
            'total_cgst': total_cgst,
            'total_sgst': total_sgst,
            'total_igst': total_igst,
            'total_amount': total_amount,
            'total_invoice': total_invoice,
            'start_date': start_date_str,
            'end_date': end_date_str,
            'total_freight':total_freight
        }
        return render(request, template_path.purchase_register_list_report, context)

    return render(request, template_path.purchase_register_date_report)



#  HSN Wise Report
@login_required
def hsn_wise_date(request):
    return render(request, template_path.hsn_wise_date_report)

@login_required
def hsn_wise_by_date_range(request):
    start_date = None
    end_date = None
    if request.method == 'POST':
        start_date_str = request.POST.get('start_date')
        end_date_str = request.POST.get('end_date')

        start_date = datetime.strptime(start_date_str, '%Y-%m-%d').date()
        end_date = datetime.strptime(end_date_str, '%Y-%m-%d').date()

        request.session['start_date'] = start_date_str
        request.session['end_date'] = end_date_str


        

        # Filter items within the date range
        items_within_range = invoice_items.objects.filter(date__range=(start_date, end_date))

        # Extract specific fields for display
        filtered_items = items_within_range.values(
            'invoice_hsn', 'invoice_description_goods',  'invoice_godown','invoice_qantity','invoice_unit_price'
        )

        # Calculate total amount
        total_amount = items_within_range.aggregate(total_amount=Sum('invoice_tax_rate'))['total_amount'] or 0

        return render(request, template_path.hsn_wise_report, {
            'filtered_items': filtered_items,
            'total_amount': total_amount ,
            'start_date': start_date,
            'end_date': end_date,
        })
   

    return render(request, template_path.hsn_wise_date_report)


@login_required
def download_hsn_wise_excel(request):
    start_date_str = request.session.get('start_date')
    end_date_str = request.session.get('end_date')

    # Validate date strings
    if not start_date_str or not end_date_str:
        return HttpResponse("Start date and end date are required.", status=400)

    try:
        start_date = datetime.strptime(start_date_str, '%Y-%m-%d').date()
        end_date = datetime.strptime(end_date_str, '%Y-%m-%d').date()
    except ValueError:
        return HttpResponse("Invalid date format. Please use YYYY-MM-DD.", status=400)

    # Retrieve data from the database based on the date range
    items_within_range = invoice_items.objects.filter(date__range=(start_date, end_date))

    # Create an in-memory output file for the new workbook
    output = BytesIO()
    workbook = openpyxl.Workbook()
    sheet = workbook.active

    # Add title and date range
    title = "M/S KEYMECH TECHNOLOGIES"
    subtitle = "HSN Wise Report"
    date_range = f"From {start_date_str} To {end_date_str}"

    sheet.merge_cells('A1:E1')
    sheet.merge_cells('A2:E2')
    sheet.merge_cells('A3:E3')

    title_cell = sheet.cell(row=1, column=1)
    subtitle_cell = sheet.cell(row=2, column=1)
    date_range_cell = sheet.cell(row=3, column=1)

    title_cell.value = title
    subtitle_cell.value = subtitle
    date_range_cell.value = date_range

    title_cell.font = Font(size=14, bold=True)
    subtitle_cell.font = Font(size=12, bold=True)
    date_range_cell.font = Font(size=10, bold=True)

    # Write headers
    headers = ['HSN Code', 'Description', 'Godown', 'Quantity', 'Unit Price']
    sheet.append([])
    sheet.append(headers)

    total_amount = 0.0  # Initialize total amount

    # Write data rows
    for item in items_within_range:
        sheet.append([
            item.invoice_hsn,
            item.invoice_description_goods,
            item.invoice_godown,
            item.invoice_qantity,
            float(item.invoice_unit_price or 0)  # Convert to float
        ])
        total_amount += float(item.invoice_tax_rate or 0)  

    # Add total amount at the end
    sheet.append([])  # Empty row for spacing
    sheet.append(['', '', '', 'Total Amount', total_amount])

    # Set column widths for better visibility
    column_widths = [15, 30, 20, 15, 15]
    for i, column_width in enumerate(column_widths, start=1):
        sheet.column_dimensions[openpyxl.utils.get_column_letter(i)].width = column_width

    # Save the workbook to the output file
    workbook.save(output)
    output.seek(0)

    # Create the response
    response = HttpResponse(
        output,
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename=hsn_wise_report.xlsx'

    return response


# def inventory_division(request):
#     return render(request, template_path.hsn_wise_date_report)
from django.db.models import Sum, Q, F, FloatField


from django.db.models.functions import Cast


@login_required
def inventory_division_date(request):
    inventory_list = inventory.objects.all()  
    return render(request, template_path.inventory_division_date_report, {'inventory_list': inventory_list})




@login_required
def display_inventory_division_date_range(request):
    inventory_name = None

    if request.method == 'POST':
        inventory_name = request.POST.get('inventory_name')

        # Filter items by inventory name
        items_within_range = inventory.objects.filter(
            inventory_name__icontains=inventory_name
        ).order_by('inventory_name')

        # Extract specific fields for display
        filtered_items = list(items_within_range.values(
            'item_code', 'inventory_name', 'units', 'opening_stock_quantity', 'purchase_rate','hsn'
        ))

        # Calculate the amount for each item and total amount
        total_amount = 0
        for item in filtered_items:
            item['amount'] = float(item['opening_stock_quantity']) * float(item['purchase_rate'])
            total_amount += item['amount']

        return render(request, template_path.inventory_division_list_report, {
            'filtered_items': filtered_items,
            'total_amount': total_amount,
            'inventory_name': inventory_name,
        })


    return render(request, template_path.inventory_division_date_report)



@login_required
def download_excel(request):
    start_date_str = request.session.get('start_date')
    end_date_str = request.session.get('end_date')

    # Validate date strings
    if not start_date_str or not end_date_str:
        return HttpResponse("Start date and end date are required.", status=400)

    try:
        start_date = datetime.strptime(start_date_str, '%Y-%m-%d').date()
        end_date = datetime.strptime(end_date_str, '%Y-%m-%d').date()
    except ValueError:
        return HttpResponse("Invalid date format. Please use YYYY-MM-DD.", status=400)

    # Retrieve data from the database based on the date range
    purchase_items_within_range = Purchase_Invoice.objects.filter(purchase_invoice_date__range=(start_date, end_date))

    # Create an in-memory output file for the new workbook
    output = BytesIO()
    workbook = openpyxl.Workbook()
    sheet = workbook.active

    # Add title and date range
    title = "M/S KEYMECH TECHNOLOGIES"
    subtitle = "Purchase Register Report"
    date_range = f"From {start_date_str} To {end_date_str}"

    sheet.merge_cells('A1:I1')
    sheet.merge_cells('A2:I2')
    sheet.merge_cells('A3:I3')

    title_cell = sheet.cell(row=1, column=1)
    subtitle_cell = sheet.cell(row=2, column=1)
    date_range_cell = sheet.cell(row=3, column=1)

    title_cell.value = title
    subtitle_cell.value = subtitle
    date_range_cell.value = date_range

    title_cell.font = Font(size=13, bold=True)
    subtitle_cell.font = Font(size=12, bold=True)
    date_range_cell.font = Font(size=10, bold=True)

    # Write headers
    headers = [
        'Vendor Name', 'Invoice No', 'Invoice Date', 'Amount', 'P&F', 'CGST', 'SGST', 'IGST', 'Total'
    ]
    sheet.append([])  # Add an empty row for spacing
    sheet.append(headers)

    total_amount = 0.0  # Initialize total amount

    # Write data rows
    for item in purchase_items_within_range:
        total = float(item.purchase_invoice_total or 0)  # Convert to float
        sheet.append([
            item.purchase_invoice_vendor_name.contact_person,
            item.purchase_invoice_grn_no,
            item.purchase_invoice_date,
            float(item.purchase_invoice_sub_total or 0),  # Convert to float
            float(item.purchase_freight_amount or 0),    # Convert to float
            float(item.purchase_invoice_cgstval or 0),    # Convert to float
            float(item.purchase_invoice_sgstval or 0),    # Convert to float
            float(item.purchase_invoice_igstval or 0),    # Convert to float
            total
        ])
        total_amount += total  # Add to total amount

    sheet.append([])  # Empty row for spacing
    sheet.append(['', '', '', '', '', '', '', 'Total', total_amount])

    # Set column widths
    column_widths = [20, 15, 15, 15, 10, 10, 10, 10, 15]
    for col_num, width in enumerate(column_widths, 1):
        column_letter = openpyxl.utils.get_column_letter(col_num)
        sheet.column_dimensions[column_letter].width = width

    # Save the workbook to the output file
    workbook.save(output)
    output.seek(0)

    # Create the response
    response = HttpResponse(
        output,
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename=purchase_register.xlsx'
    response['Total-Amount'] = total_amount  # Optionally add total amount to response headers

    return response


@login_required
def download_sales_excel(request):
    start_date_str = request.session.get('start_date')
    end_date_str = request.session.get('end_date')

    # Validate date strings
    if not start_date_str or not end_date_str:
        return HttpResponse("Start date and end date are required.", status=400)

    try:
        start_date = datetime.strptime(start_date_str, '%Y-%m-%d').date()
        end_date = datetime.strptime(end_date_str, '%Y-%m-%d').date()
    except ValueError:
        return HttpResponse("Invalid date format. Please use YYYY-MM-DD.", status=400)

    # Retrieve data from the database based on the date range
    sales_register_within_range = Invoice.objects.filter(invoice_date__range=(start_date, end_date))


    output = BytesIO()
    workbook = openpyxl.Workbook()
    sheet = workbook.active

    # Add title and date range
    title = "KEYMECH TECHNOLOGIES"
    subtitle = "Sales Register Report"
    date_range = f"From {start_date_str} To {end_date_str}"

    sheet.merge_cells('A1:K1')
    sheet.merge_cells('A2:K2')
    sheet.merge_cells('A3:K3')

    title_cell = sheet.cell(row=1, column=1)
    subtitle_cell = sheet.cell(row=2, column=1)
    date_range_cell = sheet.cell(row=3, column=1)

    title_cell.value = title
    subtitle_cell.value = subtitle
    date_range_cell.value = date_range

    title_cell.font = Font(size=14, bold=True)
    subtitle_cell.font = Font(size=12, bold=True)
    date_range_cell.font = Font(size=10, bold=True)

    # Write headers
    headers = [
        'Date', 'Customer Name', 'GST No', 'PO No', 'Invoice No', 'Invoice Date',
        'Invoice Amount', 'CGST', 'SGST', 'IGST', 'Total'
    ]
    sheet.append([])
    sheet.append(headers)

    total_amount = 0.0  # Initialize total amount

    # Write data rows
    for item in sales_register_within_range:
        total = float(item.invoice_total or 0)  # Convert to float
        sheet.append([
            item.invoice_date,
            item.invoice_customer_name.dc_customer_name.customer,
            item.invoice_gst_no,
            item.invoice_buyer_order_no,
            item.inv_number,
            item.invoice_date,
            float(item.invoice_due or 0),        # Convert to float
            float(item.invoice_cgstval or 0),    # Convert to float
            float(item.invoice_sgstval or 0),    # Convert to float
            float(item.invoice_igstval or 0),    # Convert to float
            total
        ])
        total_amount += total  # Add to total amount

    # Add total amount at the end
    sheet.append([])  # Empty row for spacing
    sheet.append(['', '', '', '', '', '', '', '', '', 'Total', total_amount])

    # Set column widths for better visibility
    column_widths = [15, 20, 20, 20, 10, 15, 20, 10, 10, 20, 15]
    for i, column_width in enumerate(column_widths, start=1):
        sheet.column_dimensions[openpyxl.utils.get_column_letter(i)].width = column_width

    # Save the workbook to the output file
    workbook.save(output)
    output.seek(0)

    # Create the response
    response = HttpResponse(
        output,
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename=sales_register.xlsx'
    response['Total-Amount'] = total_amount  # Optionally add total amount to response headers

    return response


def download_inventory_excel(request):
    start_date_str = request.session.get('start_date')
    end_date_str = request.session.get('end_date')
    selected_company = request.session.get('company_name', '').strip()

    # Validate date strings
    if not start_date_str or not end_date_str:
        return HttpResponse("Start date and end date are required.", status=400)

    try:
        start_date = datetime.strptime(start_date_str, '%Y-%m-%d').date()
        end_date = datetime.strptime(end_date_str, '%Y-%m-%d').date()
    except ValueError:
        return HttpResponse("Invalid date format. Please use YYYY-MM-DD.", status=400)

    # Retrieve data from the database based on the date range
    items_within_range = inventory.objects.filter(
        date__range=(start_date, end_date)
    )

    # Filter by selected company if provided
    if selected_company:
        items_within_range = items_within_range.filter(
            vendor_name__company_name__icontains=selected_company
        )

    items_within_range = items_within_range.order_by('date')

    # Extract fields including company name
    filtered_items = list(items_within_range.values(
        'vendor_name__company_name',
        'item_code',
        'inventory_name',
        'opening_stock_quantity',
        'opening_rate',
        'available_quantity',
        'purchase_rate'
    ))

    # Calculate amount
    total_amount = 0
    for item in filtered_items:
        try:
            qty = float(item['opening_stock_quantity'])
            rate = float(item['purchase_rate'])
        except (ValueError, TypeError):
            qty, rate = 0, 0
        item['amount'] = qty * rate
        total_amount += item['amount']

    # Create Excel
    output = BytesIO()
    workbook = openpyxl.Workbook()
    sheet = workbook.active

    # Titles
    title = "M/S KEYMECH TECHNOLOGIES"
    subtitle = "Inventory Summary Report"
    if selected_company:
        subtitle += f" - {selected_company}"
    date_range = f"From {start_date_str} To {end_date_str}"

    sheet.merge_cells('A1:F1')
    sheet.merge_cells('A2:F2')
    sheet.merge_cells('A3:F3')

    sheet.cell(row=1, column=1, value=title).font = Font(size=14, bold=True)
    sheet.cell(row=2, column=1, value=subtitle).font = Font(size=12, bold=True)
    sheet.cell(row=3, column=1, value=date_range).font = Font(size=10, bold=True)

    # Headers
    headers = ['Company Name', 'Item Code', 'MM NUMBER', 'Quantity','Opening Rate', 'Rate','Closing Balance', 'Amount']
    sheet.append([])
    sheet.append(headers)

    # Data rows
    for item in filtered_items:
        sheet.append([
            item['vendor_name__company_name'],
            item['item_code'],
            item['inventory_name'],
            item['opening_stock_quantity'],
            item['opening_rate'],
            item['purchase_rate'],
            item['available_quantity'],
            item['amount']
        ])

    # Total row
    sheet.append([])
    sheet.append(['', '', '', '','','', 'Total', total_amount])

    # Column widths
    column_widths = [30, 15, 30, 15, 15,15, 20]
    for i, width in enumerate(column_widths, start=1):
        sheet.column_dimensions[openpyxl.utils.get_column_letter(i)].width = width

    workbook.save(output)
    output.seek(0)

    response = HttpResponse(
        output,
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename=inventory_summary.xlsx'
    return response

@login_required
def purchase_register_detail(request, id):
    invoice = get_object_or_404(Purchase_Invoice, id=id)
    items = Purchase_Invoice_items.objects.filter(purchase_invoice_id=id)

    context = {
        'invoice': invoice, 
        

    }
    return render(request, 'Report/purchase_register_view.html', {'invoice': invoice, 'items': items})
    # return render(request, template_path.purchase_register_detail,context)


@login_required
def sales_display_items_by_date_range(request):
    if request.method == 'POST':
        start_date_str = request.POST.get('start_date')
        end_date_str = request.POST.get('end_date')

        start_date = datetime.strptime(start_date_str, '%Y-%m-%d').date()
        end_date = datetime.strptime(end_date_str, '%Y-%m-%d').date()

        # Filter items within the date range
        items_within_range = Invoice.objects.filter(invoice_date__range=(start_date, end_date))

        return render(request, template_path.sales_register_list_report, {'items_within_range': items_within_range})

    return render(request, template_path.sales_register_date_report)


def parse_date(date_str):
    for fmt in ('%B %d, %Y', '%Y-%m-%d'):
        try:
            return datetime.strptime(date_str, fmt).date()
        except ValueError:
            continue
    raise ValueError(f"Date format not recognized: {date_str}")

def get_last_purchase_unit(item_code, start_date, end_date):
    latest_purchase_invoice = Purchase_Invoice.objects.filter(
        id__in=Purchase_Invoice_items.objects.filter(
            purchase_invoice_item_code=item_code
        ).values_list('purchase_invoice_id', flat=True),
        purchase_invoice_date__range=(start_date, end_date)
    ).order_by('-purchase_invoice_date').first()

    if not latest_purchase_invoice:
        # print("No latest purchase invoice found.")
        return None

    # print(latest_purchase_invoice.id, "latest_purchase_invoice.id")

    purchase_invoice_items_data = Purchase_Invoice_items.objects.filter(
        purchase_invoice_item_code=item_code,
        purchase_invoice_id=latest_purchase_invoice.id
    ).last()

    if purchase_invoice_items_data:
        # print(purchase_invoice_items_data.purchase_invoice_uom, "purchase_invoice_items_data.purchase_invoice_uom")
        return purchase_invoice_items_data.purchase_invoice_uom
    else:
        # print("No purchase invoice item data found.")
        return None



def select_dates(request):
    print('dfsdgprint')
    
    return render(request, 'Report/select_dates.html')



def inventory_overview_csv(request):

    if request.method == "POST":
        start_date_str = request.POST.get('start_date')
        end_date_str = request.POST.get('end_date')
    else:
        start_date_str = request.GET.get('start_date')
        end_date_str = request.GET.get('end_date')

    download_csv = request.GET.get('download_csv')  # keep CSV download as GET param

    # Parse dates safely
    try:
        if not start_date_str or not end_date_str:
            return render(request, 'Report/select_dates.html', {
                'error_message': 'Please select both From and To dates.'
            })

        start_date = parse(start_date_str).date()
        end_date = parse(end_date_str).date()

        if start_date > end_date:
            raise ValueError("Start date after end date")
    except Exception:
        return render(request, 'Report/select_dates.html', {
            'error_message': 'Invalid start/end date. Please check again.'
        })

    combined_data = []

    all_items = inventory.objects.all()

    for item in all_items:
        item_code = item.item_code
        item_name = item.inventory_name
        uom = item.units # assumes you have a field like this
        units_per_case = item.unit_per_case   # Default to 1 if not set

        invoices = Invoice.objects.filter(
            id__in=invoice_items.objects.filter(invoice_item_code=item_code)
                .values_list('invoice_id', flat=True),
            invoice_date__range=(start_date, end_date)
        )
        sales_items = invoice_items.objects.filter(
            invoice_item_code=item_code,
            invoice_id__in=invoices.values_list('id', flat=True)
        )

        for inv in invoices:
            customer = inv.invoice_customer_name_customer
            items = sales_items.filter(invoice_id=inv.id)

            for s_item in items:
                qty = float(s_item.invoice_qantity or 0)
                is_free = (float(inv.invoice_total or 0) == 0)
                
                combined_data.append({
                    'type': 'Sales',
                    'number': inv.inv_number,
                    'Cust_vendor_code':customer.customer_code,
                    'Cust_vendor_name':customer.customer,
                    'Cust_vendor_area':customer.street,
                    'Cust_vendor_billing':customer.city,
                    'Cust_vendor_add2':customer.get_full_address_customer(),
                    'Cust_vendor_add3': customer.com_state,                    
                    'vendor_gst_no':customer.gst_number,
                    'Cust_vendor_pan_no':customer.pan_number,
                    'date': inv.invoice_date,
                    'licence_no': customer.licence_no or "",
                    'contact_no': customer.phone or '',
                    'product_code': item_code,
                    'product_name': item_name,
                    'company_name': customer.company_name,  # Or item.company_name if present
                    'uom': uom,
                    'units_per_case': units_per_case,
                    'out_quantity':qty,
                    'in_quantity': 0,
                    'out_free': qty if is_free else 0,       
                    'in_free': 0,
                    'bill_no': inv.inv_number or inv.id,
                    'bill_date': inv.invoice_date,
                })

        purchase_invoices = Purchase_Invoice.objects.filter(
            id__in=Purchase_Invoice_items.objects.filter(
                purchase_invoice_item_code=item_code
            ).values_list('purchase_invoice_id', flat=True),
            purchase_invoice_date__range=(start_date, end_date)
        )
        purchase_items = Purchase_Invoice_items.objects.filter(
            purchase_invoice_item_code=item_code,
            purchase_invoice_id__in=purchase_invoices.values_list('id', flat=True)
        )

        for p_inv in purchase_invoices:
            vendor = p_inv.purchase_invoice_vendor_name
            items = purchase_items.filter(purchase_invoice_id=p_inv.id)

            for p_item in items:
                qty = float(p_item.purchase_invoice_qantity or 0)
                is_free = (float(p_inv.purchase_invoice_total or 0) == 0)

                combined_data.append({
                    'type': 'Purchase',
                    'number': p_inv.purchase_invoice_no,
                    'date': p_inv.purchase_invoice_date,
                    'Cust_vendor_code':vendor.vendor_code,
                    'Cust_vendor_name':vendor.contact_person,
                    'Cust_vendor_area':vendor.vendor_street,
                    'Cust_vendor_billing':vendor.vendor_city,
                    'Cust_vendor_add2':vendor.get_full_address(),
                    'Cust_vendor_add3': vendor.vendor_state,                    
                    'vendor_gst_no':vendor.gst_number,
                    'Cust_vendor_pan_no':vendor.pan_number,
                    'licence_no': vendor.licence_no or ""  ,
                    'contact_no': vendor.phone or '',
                    'product_code': item_code,
                    'product_name': item_name,
                    'company_name': vendor.company_name,
                    'uom': uom,
                    'units_per_case': units_per_case,
                    'out_quantity': 0,
                    'in_quantity': qty,
                    'out_free': 0,
                    'in_free': qty if is_free else 0,
                    'bill_no': p_inv.purchase_invoice_PO_no ,
                    'bill_date': p_inv.purchase_invoice_date,
                })
    grouped_data = {}

    for row in combined_data:
        key = (row['company_name'], row['product_code'])

        if key not in grouped_data:
            grouped_data[key] = row.copy()
        else:
            grouped_data[key]['out_quantity'] += row['out_quantity']
            grouped_data[key]['in_quantity'] += row['in_quantity']
            grouped_data[key]['out_free'] += row['out_free']
            grouped_data[key]['in_free'] += row['in_free']

            # Optional: keep latest bill_date
            
            if row['bill_date'] > grouped_data[key]['bill_date']:
                grouped_data[key]['bill_date'] = row['bill_date']
            # Optional: keep latest bill_no
            grouped_data[key]['bill_no'] = row['bill_no']


    combined_data = sorted(grouped_data.values(), key=lambda x: x['bill_date'])
    # ✅ CSV Export
    if download_csv:
        response = HttpResponse(content_type='text/csv')
        response['Content-Disposition'] = 'attachment; filename=inventory_report.csv'
        writer = csv.writer(response)

        writer.writerow([
            'DOCTYPE',
            'DOCNO',
            'DOCDATE',
            'CUSTOMER /VENDOR CODE',
            'CUSTOMER /VENDOR NAME',
            'CUSTOMER /VENDOR AREA',
            'CUSTOMER /VENDOR BILLING',
            'CUSTOMER /VENDOR ADD2',
            'CUSTOMER /VENDOR ADD3',
            'CUSTOMER /VENDOR GST NO',
            'CUSTOMER /VENDOR PAN NO',
            'CUSTOMER /VENDOR Licence No.',
            'CUSTOMER / VENDOR Contact No',
            'PRODUCTCODE',
            'PRODUCTNAME',
            'COMPANYNAME',
            'UOM',
            'UNITSPERCASE',
            'OUTQUANTITY(UNITS)',
            'INQUANTITY(UNITS)',
            'OUTFREE(UNITS)',
            'INFREE(UNITS)',
            'CHALAN / VENDOR BILL NO',
            'CHALAN / VENDOR BILL DATE',
        ])

        for row in combined_data:
            writer.writerow([
                row['type'],row['number'],row['date'], row['Cust_vendor_code'], row['Cust_vendor_name'], row['Cust_vendor_area'],
                row['Cust_vendor_billing'],row['Cust_vendor_add2'], row['Cust_vendor_add3'], row['vendor_gst_no'], row['Cust_vendor_pan_no'],
                row['licence_no'], row['contact_no'], row['product_code'], row['product_name'],
                row['company_name'], row['uom'], row['units_per_case'],
                row['out_quantity'], row['in_quantity'], row['out_free'], row['in_free'],
                row['bill_no'], row['bill_date'],
            ])
        return response

    context = {
        'combined_data': combined_data,
        'start_date': start_date,
        'end_date': end_date,
    }
    return render(request, 'Report/inventory_entity_data.html', context)
